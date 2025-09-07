import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_matching_entities_spreadsheet():
    # Create workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    section_font = Font(bold=True, size=11, color="FFFFFF")
    core_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    matching_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    missing_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    
    subheader_font = Font(bold=True, italic=True)
    regular_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # 1. BOOKS SHEET
    books_sheet = wb.create_sheet("Books", 0)
    books_data = {
        "headers": ["Field Name", "Data Type", "Description", "Current Status", "Notes"],
        "core": [
            ["title", "VARCHAR(500)", "Book title", "✓ Exists", "Required field"],
            ["author", "VARCHAR(255)", "Author name(s)", "✓ Exists", "Required field"],
            ["description", "TEXT", "Book description/summary", "✓ Exists", "Long-form description"],
            ["cover_image_url", "VARCHAR(500)", "URL to book cover image", "✓ Exists", "For visual display"],
            ["amazon_url", "VARCHAR(500)", "Amazon purchase link", "✓ Exists", "For purchasing"],
            ["publication_year", "INTEGER", "Year published", "✓ Exists", "For relevance"],
            ["reading_time", "VARCHAR(100)", "Estimated time to read", "✓ Exists", "e.g., '4-5 hours'"],
            ["difficulty_level", "VARCHAR(50)", "Reading difficulty", "✓ Exists", "Beginner/Intermediate/Advanced"],
            ["is_active", "BOOLEAN", "Active status", "✓ Exists", "Show/hide from matching"],
            ["created_at", "TIMESTAMP", "Record creation date", "✓ Exists", "Audit trail"],
            ["updated_at", "TIMESTAMP", "Last update date", "✓ Exists", "Audit trail"]
        ],
        "matching": [
            ["focus_areas_covered", "TEXT (JSON)", "Business focus areas covered", "✓ Exists", "JSON array for matching"],
            ["topics", "TEXT (JSON)", "Topics covered in book", "✓ Exists", "JSON array of topics"],
            ["target_audience", "TEXT", "Ideal reader description", "✓ Exists", "Text description"],
            ["key_takeaways", "TEXT (JSON)", "Main lessons/takeaways", "✓ Exists", "JSON array of takeaways"]
        ],
        "missing": [
            ["price_range", "VARCHAR(50)", "Price bracket", "❌ Missing", "$10-20, $20-30, etc."],
            ["book_format", "TEXT (JSON)", "Available formats", "❌ Missing", "Hardcover, Paperback, eBook, Audiobook"],
            ["page_count", "INTEGER", "Number of pages", "❌ Missing", "For time estimation"],
            ["isbn", "VARCHAR(20)", "ISBN number", "❌ Missing", "Unique identifier"],
            ["publisher", "VARCHAR(255)", "Publishing company", "❌ Missing", "For credibility"],
            ["endorsements", "TEXT (JSON)", "Notable endorsements", "❌ Missing", "Industry leaders who recommend"],
            ["companion_resources", "TEXT (JSON)", "Additional resources", "❌ Missing", "Workbooks, courses, etc."],
            ["revenue_tier_relevance", "TEXT (JSON)", "Best for which contractor sizes", "❌ Missing", "Match to contractor revenue"],
            ["powerconfidence_score", "INTEGER", "Book quality rating", "❌ Missing", "0-100 score based on reviews"],
            ["total_reviews", "INTEGER", "Number of reviews", "❌ Missing", "From Amazon, Goodreads, etc."],
            ["average_rating", "DECIMAL(2,1)", "Average star rating", "❌ Missing", "1.0-5.0 scale"]
        ]
    }
    
    # 2. PODCASTS SHEET
    podcasts_sheet = wb.create_sheet("Podcasts", 1)
    podcasts_data = {
        "headers": ["Field Name", "Data Type", "Description", "Current Status", "Notes"],
        "core": [
            ["title", "VARCHAR(255)", "Podcast title", "✓ Exists", "Required field"],
            ["host", "VARCHAR(255)", "Host name(s)", "✓ Exists", "Required field"],
            ["description", "TEXT", "Podcast description", "✓ Exists", "Long-form description"],
            ["frequency", "VARCHAR(100)", "Release frequency", "✓ Exists", "Weekly, Bi-weekly, Monthly"],
            ["website", "VARCHAR(500)", "Podcast website", "✓ Exists", "Official website"],
            ["logo_url", "VARCHAR(500)", "Podcast logo/artwork", "✓ Exists", "For visual display"],
            ["is_active", "BOOLEAN", "Active status", "✓ Exists", "Show/hide from matching"]
        ],
        "matching": [
            ["focus_areas_covered", "TEXT (JSON)", "Business focus areas covered", "✓ Exists", "JSON array for matching"],
            ["topics", "TEXT (JSON)", "Topics discussed", "✓ Exists", "JSON array of topics"]
        ],
        "missing": [
            ["target_audience", "TEXT", "Ideal listener description", "❌ Missing", "Critical for matching"],
            ["episode_count", "INTEGER", "Total episodes published", "❌ Missing", "Shows longevity"],
            ["average_episode_length", "VARCHAR(50)", "Typical episode duration", "❌ Missing", "15-30 min, 30-60 min, etc."],
            ["spotify_url", "VARCHAR(500)", "Spotify podcast link", "❌ Missing", "For listening"],
            ["apple_podcasts_url", "VARCHAR(500)", "Apple Podcasts link", "❌ Missing", "For listening"],
            ["youtube_url", "VARCHAR(500)", "YouTube channel link", "❌ Missing", "If video podcast"],
            ["guest_profile", "TEXT", "Types of guests featured", "❌ Missing", "Industry leaders, contractors, etc."],
            ["format_type", "VARCHAR(100)", "Podcast format", "❌ Missing", "Interview, Solo, Panel, Q&A"],
            ["revenue_tier_relevance", "TEXT (JSON)", "Best for which contractor sizes", "❌ Missing", "Match to contractor revenue"],
            ["subscription_required", "BOOLEAN", "Requires paid subscription", "❌ Missing", "Free vs Premium"],
            ["launch_year", "INTEGER", "Year podcast started", "❌ Missing", "For credibility"],
            ["download_count", "INTEGER", "Total downloads/listens", "❌ Missing", "Popularity metric"],
            ["average_rating", "DECIMAL(2,1)", "Average listener rating", "❌ Missing", "1.0-5.0 scale"],
            ["notable_episodes", "TEXT (JSON)", "Must-listen episodes", "❌ Missing", "Top recommended episodes"]
        ]
    }
    
    # 3. EVENTS SHEET
    events_sheet = wb.create_sheet("Events", 2)
    events_data = {
        "headers": ["Field Name", "Data Type", "Description", "Current Status", "Notes"],
        "core": [
            ["name", "VARCHAR(255)", "Event name", "✓ Exists", "Required field"],
            ["date", "DATE", "Event date(s)", "✓ Exists", "When event occurs"],
            ["location", "VARCHAR(255)", "Event location", "✓ Exists", "City, State or Virtual"],
            ["format", "VARCHAR(100)", "Event format", "✓ Exists", "In-Person, Virtual, Hybrid"],
            ["description", "TEXT", "Event description", "✓ Exists", "Long-form description"],
            ["website", "VARCHAR(500)", "Event website", "✓ Exists", "Registration/info site"],
            ["logo_url", "VARCHAR(500)", "Event logo/banner", "✓ Exists", "For visual display"],
            ["registration_deadline", "DATE", "Last day to register", "✓ Exists", "Cutoff date"],
            ["expected_attendees", "VARCHAR(255)", "Expected attendance", "✓ Exists", "Size of event"],
            ["is_active", "BOOLEAN", "Active status", "✓ Exists", "Show/hide from matching"]
        ],
        "matching": [
            ["focus_areas_covered", "TEXT (JSON)", "Business focus areas covered", "✓ Exists", "JSON array for matching"]
        ],
        "missing": [
            ["target_audience", "TEXT", "Ideal attendee description", "❌ Missing", "Critical for matching"],
            ["topics", "TEXT (JSON)", "Topics covered at event", "❌ Missing", "Sessions, workshops, keynotes"],
            ["price_range", "VARCHAR(100)", "Registration cost", "❌ Missing", "$500-1000, $1000-2000, etc."],
            ["early_bird_deadline", "DATE", "Early registration deadline", "❌ Missing", "For discount pricing"],
            ["early_bird_discount", "VARCHAR(100)", "Early bird pricing/savings", "❌ Missing", "20% off, Save $200, etc."],
            ["sponsor_companies", "TEXT (JSON)", "Event sponsors", "❌ Missing", "Major sponsors/partners"],
            ["speaker_highlights", "TEXT (JSON)", "Notable speakers", "❌ Missing", "Keynote speakers, industry leaders"],
            ["ceu_credits_available", "BOOLEAN", "Continuing education credits", "❌ Missing", "Professional development"],
            ["ceu_credit_hours", "DECIMAL(3,1)", "Number of CEU hours", "❌ Missing", "If applicable"],
            ["meals_included", "VARCHAR(255)", "Meal provisions", "❌ Missing", "Breakfast, Lunch, Dinner, etc."],
            ["accommodation_info", "TEXT", "Hotel/lodging information", "❌ Missing", "Partner hotels, discounts"],
            ["revenue_tier_relevance", "TEXT (JSON)", "Best for which contractor sizes", "❌ Missing", "Match to contractor revenue"],
            ["past_attendance_count", "INTEGER", "Previous year attendance", "❌ Missing", "Historical data"],
            ["vendor_booth_available", "BOOLEAN", "Exhibitor opportunities", "❌ Missing", "Can contractors exhibit"],
            ["networking_events", "TEXT (JSON)", "Networking opportunities", "❌ Missing", "Mixers, dinners, golf, etc."],
            ["recording_available", "BOOLEAN", "Sessions recorded", "❌ Missing", "Post-event access"],
            ["refund_policy", "TEXT", "Cancellation/refund policy", "❌ Missing", "Important for planning"]
        ]
    }
    
    # 4. STRATEGIC PARTNERS SHEET (for reference)
    partners_sheet = wb.create_sheet("Strategic Partners", 3)
    partners_data = {
        "headers": ["Field Name", "Data Type", "Description", "Current Status", "Notes"],
        "core": [
            ["company_name", "VARCHAR(255)", "Company name", "✓ Exists", "Required field"],
            ["description", "TEXT", "Company description", "✓ Exists", "Long-form description"],
            ["logo_url", "VARCHAR(500)", "Company logo", "✓ Exists", "For visual display"],
            ["website", "VARCHAR(500)", "Company website", "✓ Exists", "Official website"],
            ["contact_email", "VARCHAR(255)", "Primary contact email", "✓ Exists", "Main contact"],
            ["contact_phone", "VARCHAR(50)", "Primary contact phone", "✓ Exists", "Main contact"],
            ["power100_subdomain", "VARCHAR(255)", "Power100 subdomain", "✓ Exists", "partner.power100.io"],
            ["value_proposition", "TEXT", "Value prop/tagline", "✓ Exists", "Elevator pitch"],
            ["power_confidence_score", "INTEGER", "PowerConfidence rating", "✓ Exists", "0-100 score"],
            ["is_active", "BOOLEAN", "Active status", "✓ Exists", "Show/hide from matching"],
            ["created_at", "TIMESTAMP", "Record creation date", "✓ Exists", "Audit trail"],
            ["updated_at", "TIMESTAMP", "Last update date", "✓ Exists", "Audit trail"]
        ],
        "matching": [
            ["focus_areas_served", "TEXT (JSON)", "Focus areas they serve", "✓ Exists", "JSON array for matching"],
            ["service_areas", "TEXT (JSON)", "Services offered", "✓ Exists", "JSON array of services"],
            ["target_revenue_range", "TEXT (JSON)", "Client revenue targets", "✓ Exists", "JSON array of ranges"],
            ["geographic_regions", "TEXT (JSON)", "Regions served", "✓ Exists", "JSON array of regions"],
            ["service_category", "VARCHAR(255)", "Primary service category", "✓ Exists", "Main category"],
            ["focus_areas_12_months", "TEXT (JSON)", "12-month focus areas", "✓ Exists", "Current priorities"]
        ],
        "missing": [
            ["minimum_engagement_cost", "VARCHAR(100)", "Minimum project cost", "❌ Missing", "Entry price point"],
            ["typical_engagement_duration", "VARCHAR(100)", "Typical project length", "❌ Missing", "3-6 months, 6-12 months, etc."],
            ["client_success_rate", "DECIMAL(3,1)", "Success percentage", "❌ Missing", "Based on outcomes"],
            ["total_contractors_served", "INTEGER", "Number of clients served", "❌ Missing", "Experience metric"],
            ["years_in_business", "INTEGER", "Company age", "❌ Missing", "Credibility factor"],
            ["team_size", "INTEGER", "Number of employees", "❌ Missing", "Company scale"],
            ["certifications", "TEXT (JSON)", "Industry certifications", "❌ Missing", "Credentials"],
            ["case_study_count", "INTEGER", "Published case studies", "❌ Missing", "Proof points"],
            ["response_time_hours", "INTEGER", "Typical response time", "❌ Missing", "Service level"]
        ]
    }
    
    # Function to write data to sheet
    def write_sheet_data(sheet, data):
        # Write title
        sheet['A1'] = f"{sheet.title} Entity - Database Schema & Matching Fields"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:E1')
        
        # Write headers
        row = 3
        for col, header in enumerate(data['headers'], 1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = regular_border
            cell.alignment = Alignment(horizontal='center')
        
        row += 2
        
        # Core Fields Section
        sheet.cell(row=row, column=1, value="CORE FIELDS").font = section_font
        sheet.cell(row=row, column=1).fill = core_fill
        sheet.merge_cells(f'A{row}:E{row}')
        row += 1
        
        for field_data in data['core']:
            for col, value in enumerate(field_data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.border = regular_border
                if col == 1:
                    cell.font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Matching Fields Section
        sheet.cell(row=row, column=1, value="MATCHING FIELDS").font = section_font
        sheet.cell(row=row, column=1).fill = matching_fill
        sheet.merge_cells(f'A{row}:E{row}')
        row += 1
        
        for field_data in data['matching']:
            for col, value in enumerate(field_data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.border = regular_border
                if col == 1:
                    cell.font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Missing Fields Section
        sheet.cell(row=row, column=1, value="MISSING FIELDS (Recommended for Better Matching)").font = section_font
        sheet.cell(row=row, column=1).fill = missing_fill
        sheet.merge_cells(f'A{row}:E{row}')
        row += 1
        
        for field_data in data['missing']:
            for col, value in enumerate(field_data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.border = regular_border
                if col == 1:
                    cell.font = Font(bold=True)
            row += 1
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 35
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 40
    
    # Write data to each sheet
    write_sheet_data(books_sheet, books_data)
    write_sheet_data(podcasts_sheet, podcasts_data)
    write_sheet_data(events_sheet, events_data)
    write_sheet_data(partners_sheet, partners_data)
    
    # 5. SUMMARY SHEET
    summary_sheet = wb.create_sheet("Summary", 0)
    summary_sheet['A1'] = "TPE Matching Entities - Database Schema Summary"
    summary_sheet['A1'].font = Font(bold=True, size=16)
    summary_sheet.merge_cells('A1:D1')
    
    summary_data = [
        ["", "", "", ""],
        ["Entity", "Core Fields", "Matching Fields", "Missing Fields"],
        ["Books", 11, 4, 11],
        ["Podcasts", 7, 2, 14],
        ["Events", 10, 1, 17],
        ["Strategic Partners", 12, 6, 9],
        ["", "", "", ""],
        ["MATCHING ALGORITHM NOTES:", "", "", ""],
        ["1. All entities use 'focus_areas_covered' for primary matching", "", "", ""],
        ["2. 'topics' field provides secondary matching capability", "", "", ""],
        ["3. 'target_audience' critical for relevance scoring", "", "", ""],
        ["4. 'revenue_tier_relevance' ensures size-appropriate matches", "", "", ""],
        ["", "", "", ""],
        ["PRIORITY ADDITIONS:", "", "", ""],
        ["• Podcasts: Add 'target_audience' field (CRITICAL)", "", "", ""],
        ["• Events: Add 'target_audience' and 'topics' fields (CRITICAL)", "", "", ""],
        ["• All: Add 'revenue_tier_relevance' for better contractor matching", "", "", ""],
        ["• All: Add quality/rating metrics for ranking results", "", "", ""],
        ["", "", "", ""],
        ["Generated: " + datetime.now().strftime("%Y-%m-%d %H:%M"), "", "", ""]
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_sheet.cell(row=row_idx + 2, column=col_idx, value=value)
            if row_idx == 1:  # Headers
                cell.font = header_font
                cell.fill = header_fill
                cell.border = regular_border
            elif row_idx in [2, 3, 4, 5]:  # Data rows
                cell.border = regular_border
                if col_idx == 1:
                    cell.font = Font(bold=True)
            elif "NOTES:" in str(value) or "ADDITIONS:" in str(value):
                cell.font = Font(bold=True, size=12)
    
    # Adjust column widths for summary
    summary_sheet.column_dimensions['A'].width = 50
    summary_sheet.column_dimensions['B'].width = 15
    summary_sheet.column_dimensions['C'].width = 15
    summary_sheet.column_dimensions['D'].width = 15
    
    # Save the workbook
    filename = "TPE_Matching_Entities_Schema.xlsx"
    wb.save(filename)
    print(f"Spreadsheet created successfully: {filename}")
    return filename

# Run the function
if __name__ == "__main__":
    create_matching_entities_spreadsheet()